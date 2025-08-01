#!python3
# -*- coding: utf8 -*-

import csv
import os
import random
import string
import requests
import logging
from datetime import datetime
import urllib3
from bs4 import BeautifulSoup
from requests.auth import HTTPBasicAuth
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

try:
    from credentials import WEB_USERNAME, WEB_PASSWORD, BASIC_AUTH_USER, BASIC_AUTH_PASS
except ImportError:
    logging.error("Файл credentials.py не найден или не содержит необходимых учетных данных")
    exit(1)

# Отключаем предупреждения о недоверенных SSL сертификатах
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

   # Настройка
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('mailbox_creation.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# Конфигурация
BASE_URL = 'https://mx.zhky.ru/mxadm'
LOGIN_URL = f'{BASE_URL}/login.php'
EDIT_URL = f'{BASE_URL}/edit.php?table=mailbox'
CREATED_ACCOUNTS_FILE = 'created_accounts.xlsx'  # Изменено на Excel формат
DEFAULT_QUOTA = '2000'  # kvota
DEFAULT_OTHER_EMAIL = 'it@vmf.zhky.ru' # pochta dlia vosstanovkenia parilia
DOMAIN = 'vmf.zhky.ru'
MAX_RETRIES = 3  # popitki esli pochta ne sozdaectia
RETRY_DELAY = 2 # vremia ojhidania


def init_excel_file():
    """Инициализирует Excel файл с заголовками"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Созданные аккаунты"

    # Заголовки столбцов
    headers = ["ФИО", "Email", "Пароль", "Дата создания", "Статус"]

    # Записываем заголовки с форматированием
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 30  # ФИО
    ws.column_dimensions['B'].width = 25  # Email
    ws.column_dimensions['C'].width = 15  # Пароль
    ws.column_dimensions['D'].width = 20  # Дата создания
    ws.column_dimensions['E'].width = 20  # Статус

    wb.save(CREATED_ACCOUNTS_FILE)


def save_to_excel(data):
    """Сохраняетв Excel файл"""
    try:
        from openpyxl import load_workbook

        if not os.path.exists(CREATED_ACCOUNTS_FILE):
            init_excel_file()

        wb = load_workbook(CREATED_ACCOUNTS_FILE)
        ws = wb.active

        next_row = ws.max_row + 1

        # Записываем данные
        ws.cell(row=next_row, column=1, value=data['name'])
        ws.cell(row=next_row, column=2, value=data['email'])
        ws.cell(row=next_row, column=3, value=data['password'])
        ws.cell(row=next_row, column=4, value=data['created_time'])
        ws.cell(row=next_row, column=5, value=data['status'])

        wb.save(CREATED_ACCOUNTS_FILE)
        return True

    except Exception as e:
        logging.error(f"Ошибка при сохранении в Excel: {str(e)}")
        return False


def find_csv_file():
    """pervii nah CSV fail"""
    for file in os.listdir('.'):
        if file.endswith('.csv') and file != 'created_accounts.csv':
            return file
    return None


def get_headers():
    """Возвращает стандартные заголовки для запросов"""
    return {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'ru,en;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Host': 'mx.zhky.ru',
        'Origin': 'https://mx.zhky.ru',
        'Referer': LOGIN_URL,
        'Sec-Ch-Ua': '"Chromium";v="124", "YaBrowser";v="24", "Not-A.Brand";v="99"',
        'Sec-Ch-Ua-Mobile': '?0',
        'Sec-Ch-Ua-Platform': '"Windows"',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.6367.770 YaBrowser/24.6.3.770 Yowser/2.5 Safari/537.36',
    }


def get_session_data():
    """Получает cookies и токен через двойную авторизацию"""
    try:
        session = requests.Session()
        headers = get_headers()

        session.auth = HTTPBasicAuth(BASIC_AUTH_USER, BASIC_AUTH_PASS)

        response = session.get(
            LOGIN_URL,
            headers=headers,
            verify=False,
            timeout=30
        )

        if response.status_code != 200:
            logging.error(f"Ошибка при получении страницы входа: {response.status_code}")
            return None, None

        soup = BeautifulSoup(response.text, 'html.parser')
        token = soup.find('input', {'name': 'token'}).get('value', '')

        if not token:
            logging.error("Не удалось найти CSRF-токен на странице входа")
            return None, None

        login_data = {
            'fUsername': WEB_USERNAME,
            'fPassword': WEB_PASSWORD,
            'lang': 'ru',
            'submit': 'Вход',
            'token': token
        }

        auth_response = session.post(
            LOGIN_URL,
            data=login_data,
            headers=headers,
            verify=False,
            allow_redirects=True,
            timeout=30
        )

        edit_response = session.get(
            EDIT_URL,
            headers=headers,
            verify=False,
            timeout=30
        )

        if edit_response.status_code != 200:
            logging.error(f"Ошибка при получении формы создания: {edit_response.status_code}")
            return None, None

        edit_soup = BeautifulSoup(edit_response.text, 'html.parser')
        edit_token = edit_soup.find('input', {'name': 'token'}).get('value', '')

        if not edit_token:
            logging.error("Не удалось найти токен в форме создания ящика")
            return None, None

        return session.cookies.get_dict(), edit_token

    except Exception as e:
        logging.error(f"Ошибка при получении сессии: {str(e)}")
        return None, None


def generate_password(length=9):
    """Генерация случайного пароля заданной длины"""
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))


def create_mailbox_with_retry(name, email, cookies, token):
    """Создание почтового ящика с попытками"""
    attempts = 0
    last_error = ""

    while attempts < MAX_RETRIES:
        attempts += 1
        success, result = create_mailbox(name, email, cookies, token)

        if success:
            return True, result

        last_error = result
        logging.warning(f"Попытка {attempts} из {MAX_RETRIES} не удалась для {name}: {result}")

        if attempts < MAX_RETRIES:
            time.sleep(RETRY_DELAY)

    return False, last_error


def create_mailbox(name, email, cookies, token):
    """Создание почтового ящика через POST запрос"""
    try:
        local_part = email.split('@')[0]
        password = generate_password()

        post_data = {
            'table': 'mailbox',
            'token': token,
            'value[local_part]': local_part,
            'value[domain]': DOMAIN,
            'value[password]': password,
            'value[password2]': password,
            'value[name]': name,
            'value[quota]': DEFAULT_QUOTA,
            'value[active]': '1',
            'value[welcome_mail]': '1',
            'value[email_other]': DEFAULT_OTHER_EMAIL,
            'submit': 'Создать ящик'
        }

        headers = get_headers()
        headers['Content-Length'] = str(len(post_data))
        headers['Referer'] = EDIT_URL

        auth = HTTPBasicAuth(BASIC_AUTH_USER, BASIC_AUTH_PASS)

        response = requests.post(
            EDIT_URL,
            data=post_data,
            cookies=cookies,
            headers=headers,
            auth=auth,
            verify=False,
            timeout=30
        )

        time.sleep(1)
        if response.status_code == 200:
            if "успешно" in response.text.lower() or "success" in response.text.lower():
                return True, password
            return False, "Создание не подтверждено в ответе сервера"
        elif response.status_code == 401:
            new_cookies, new_token = get_session_data()
            if new_cookies and new_token:
                cookies.update(new_cookies)
                token = new_token
                return False, "Сессия обновлена, попробуйте снова"
            return False, "Ошибка 401: Неавторизованный доступ"
        else:
            return False, f"HTTP ошибка: {response.status_code}"

    except Exception as e:
        return False, str(e)


def process_csv_file(csv_file, cookies, token):
    """Обработка CSV файла и создание почтовых ящиков"""
    success_count = 0
    fail_count = 0

    try:
        with open(csv_file, 'r', encoding='cp1251') as f:
            reader = csv.reader(f, delimiter=';')

            for row in reader:
                if len(row) < 2:
                    continue

                name = row[0].strip()
                email = row[1].strip()

                logging.info(f"Обработка записи: {name}, email: {email}")

                if not email or '@' not in email:
                    error_msg = f"Неверный email: {email}"
                    logging.error(error_msg)
                    save_to_excel({
                        'name': name,
                        'email': email,
                        'password': '',
                        'created_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'status': error_msg
                    })
                    fail_count += 1
                    continue

                success, result = create_mailbox_with_retry(name, email, cookies, token)
                created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                if success:
                    logging.info(f"Успешно создан ящик {email} для {name}")
                    save_to_excel({
                        'name': name,
                        'email': email.split('@')[0] + '@vmf.zhky.ru',
                        'password': result,
                        'created_time': created_time,
                        'status': 'Успешно создан'
                    })
                    success_count += 1
                else:
                    error_msg = result
                    logging.error(f"Ошибка при создании ящика для {name}: {error_msg}")
                    save_to_excel({
                        'name': name,
                        'email': email,
                        'password': '',
                        'created_time': created_time,
                        'status': error_msg
                    })
                    fail_count += 1

    except Exception as e:
        logging.error(f"Ошибка при обработке CSV файла: {str(e)}")
        return 0, 0

    return success_count, fail_count


def main():
    # Инициализируем Excel файл
    if not os.path.exists(CREATED_ACCOUNTS_FILE):
        init_excel_file()

    # Находим CSV файл
    csv_file = find_csv_file()
    if not csv_file:
        logging.error("Не найден CSV файл в текущей директории")
        return

    logging.info(f"Найден CSV файл: {csv_file}")

    # Получаем сессию
    logging.info("Получение сессионных данных...")
    cookies, token = get_session_data()

    if not cookies or not token:
        logging.error("Не удалось получить сессию. Проверьте учетные данные.")
        return

    logging.info(f"Начало обработки файла: {csv_file}")
    start_time = datetime.now()

    success_count, fail_count = process_csv_file(csv_file, cookies, token)

    end_time = datetime.now()
    duration = end_time - start_time

    logging.info("\n=== ИТОГОВЫЙ ОТЧЕТ ===")
    logging.info(f"Всего обработано записей: {success_count + fail_count}")
    logging.info(f"Успешно создано ящиков: {success_count}")
    logging.info(f"Не удалось создать ящиков: {fail_count}")
    logging.info(f"Время выполнения: {duration}")
    logging.info(f"\nРезультаты сохранены в файл: {CREATED_ACCOUNTS_FILE}")
    logging.info("Обработка завершена")


if __name__ == '__main__':
    # Проверяем наличие необходимых библиотек
    try:
        import openpyxl
    except ImportError:
        print("Ошибка: Необходимо установить библиотеку openpyxl")
        print("Установите ее командой: pip install openpyxl")
        exit(1)


    main()